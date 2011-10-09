"""
House Roll Call Votes to Excel
"""

__author__ = "Michael Myers <michael.morris.myers@gmail.com>"
__version__ = 1.0
__license__ = "MIT"

import urllib2

from datetime import datetime
from BeautifulSoup import BeautifulSoup, BeautifulStoneSoup
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter


def roll_call_to_xlsx(year, minRC = 1, maxRC = 1):
    """
    Get roll call data for a given year and stores in 
    """
    congressData = dict()
    nameList = list()
    voteData = dict()

    #Check to see if proper arguments passed in function
    if year not in range(1990, datetime.now().year+1):
        print "Date out of range, must be between 1990 and current year"
        return

    #No value was entered, default to max
    if maxRC == 1:
        maxRC = get_max_roll_call(year)

    if minRC > maxRC:
        print "Max is less than min"
        return

    for rollcall in range(minRC, maxRC+1):

        voteData.clear()

        rollcallstr = str(rollcall)
        url = 'http://clerk.house.gov/evs/' + str(year) + '/roll' + rollcallstr.zfill(3) + '.xml'

        page = urllib2.urlopen(url)
        soup = BeautifulStoneSoup(page)

        votes = soup.findAll('recorded-vote')

        numVotes = len(votes)

        for i in range(0, numVotes):
            name = votes[i].contents[0].contents[0]
            rawVote = votes[i].contents[1].contents[0]

            if 'Yes' in rawVote or 'Aye' in rawVote or 'Yea' in rawVote:
                vote = 'Yes'

            elif 'Not Voting' in rawVote:
                vote = 'Not Voting'

            elif 'Present' in rawVote:
                vote = 'Present'

            elif 'No' in rawVote or 'Nay' in rawVote:
                vote = 'No'

            else:
                vote = rawVote

            if name not in nameList:
                nameList.append(name)

            if name not in voteData:
                voteData[name] = vote

        if rollcall not in congressData:
            congressData[rollcall] = voteData.copy()
            print 'Added voteData for rollcall ' , rollcall

    nameList.sort()

    #Data received, now time for excel
    
    wb = Workbook()
    ew = ExcelWriter(workbook = wb)
    
    dest_filename = str(get_congress(year)) +'_Congress_Roll_Call_Data.xlsx'

    ws = wb.worksheets[0]
    ws.title = str(get_congress(year)) + "Congress Roll Call Data"

    colcount = len(congressData) 
    rowcount = len(nameList)

    #Write roll call numbers as headers
    for col in range(0, colcount):
        ws.cell(row= 0, column= col + 1).value = minRC + col

    #Write names in first column
    for rw in range(rowcount):
        ws.cell( row = rw + 1, column = 0 ).value = nameList[rw]

    #Write the vote data for each roll call
    i = 0
    for col in range(minRC, minRC + colcount):    #iterate through each column
        vD = congressData[col]        #get the vote data for this rollcall
        print col
        i = i + 1
        for rw in range(0, rowcount): #iterate through the rows
            if nameList[rw] in vD.keys():
                ws.cell( row = rw + 1, column = i ).value = vD[nameList[rw]]                
    ew.save( filename = dest_filename )

    return

def get_max_roll_call(year):
    """Get the total number of Roll Calls for a given year of Congress
    Returns Integer
    """
    
    url = 'http://clerk.house.gov/evs/' + str(year) + '/index.asp'
    page = urllib2.urlopen(url)
    soup = BeautifulSoup(page)    
    text = soup.find('a')

    return int(text.contents[0])

def get_congress(year):
    """Returns the Congress number for a given year
    Borrowed from python-nytcongress / nytcongress.py by Chris Amico
    https://github.com/eyeseast/python-nytcongress
    """

    return (year - 1789) / 2 + 1


